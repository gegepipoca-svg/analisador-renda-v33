"""
═══════════════════════════════════════════════════════════════
    ANALISADOR DE RENDA V3.4 - STREAMLIT
    Sistema de Análise de Extratos Bancários
    DESIGN PROFISSIONAL + DASHBOARD VISUAL
═══════════════════════════════════════════════════════════════
"""

import streamlit as st
import anthropic
import os
import base64
import json
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import PyPDF2
import io
from PIL import Image
import time
import re
import plotly.graph_objects as go
import plotly.express as px
from collections import defaultdict

# ═══════════════════════════════════════════════════════════════
# CONFIGURAÇÕES
# ═══════════════════════════════════════════════════════════════

st.set_page_config(
    page_title="Analisador de Renda Pro",
    page_icon="💰",
    layout="wide",
    initial_sidebar_state="expanded"
)

# Estilo CSS PROFISSIONAL com gradientes e animações
st.markdown("""
<style>
    /* RESET E BASE */
    .main {
        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
        background-attachment: fixed;
    }
    
    .block-container {
        padding-top: 2rem !important;
        padding-bottom: 2rem !important;
        max-width: 1200px;
    }
    
    /* HEADER IMPACTANTE */
    .hero-header {
        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
        padding: 3rem 2rem;
        border-radius: 20px;
        text-align: center;
        margin-bottom: 2rem;
        box-shadow: 0 20px 60px rgba(0,0,0,0.3);
        animation: fadeInDown 0.8s ease-out;
    }
    
    .hero-title {
        font-size: 3.5rem;
        font-weight: 900;
        background: linear-gradient(135deg, #fff 0%, #f0f0f0 100%);
        -webkit-background-clip: text;
        -webkit-text-fill-color: transparent;
        margin-bottom: 0.5rem;
        text-shadow: 2px 2px 4px rgba(0,0,0,0.2);
    }
    
    .hero-subtitle {
        font-size: 1.3rem;
        color: rgba(255,255,255,0.9);
        margin-top: 0.5rem;
        font-weight: 300;
    }
    
    .hero-icon {
        font-size: 4rem;
        margin-bottom: 1rem;
        animation: bounce 2s infinite;
    }
    
    /* CARDS DE FEATURES */
    .feature-card {
        background: white;
        padding: 1.5rem;
        border-radius: 15px;
        box-shadow: 0 10px 30px rgba(0,0,0,0.1);
        margin-bottom: 1rem;
        transition: all 0.3s ease;
        border-left: 5px solid #667eea;
    }
    
    .feature-card:hover {
        transform: translateY(-5px);
        box-shadow: 0 15px 40px rgba(0,0,0,0.2);
    }
    
    .feature-icon {
        font-size: 2.5rem;
        margin-bottom: 0.5rem;
    }
    
    .feature-title {
        font-size: 1.2rem;
        font-weight: bold;
        color: #333;
        margin-bottom: 0.5rem;
    }
    
    .feature-desc {
        color: #666;
        font-size: 0.95rem;
    }
    
    /* DASHBOARD DE RESULTADOS */
    .metric-card {
        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
        padding: 2rem;
        border-radius: 20px;
        text-align: center;
        color: white;
        box-shadow: 0 15px 35px rgba(102, 126, 234, 0.4);
        transition: all 0.3s ease;
        margin-bottom: 1.5rem;
    }
    
    .metric-card:hover {
        transform: translateY(-8px);
        box-shadow: 0 20px 45px rgba(102, 126, 234, 0.5);
    }
    
    .metric-icon {
        font-size: 3rem;
        margin-bottom: 1rem;
        opacity: 0.9;
    }
    
    .metric-label {
        font-size: 1rem;
        text-transform: uppercase;
        letter-spacing: 2px;
        opacity: 0.9;
        margin-bottom: 0.5rem;
    }
    
    .metric-value {
        font-size: 2.8rem;
        font-weight: 900;
        margin-bottom: 0.5rem;
        text-shadow: 2px 2px 4px rgba(0,0,0,0.2);
    }
    
    .metric-subtitle {
        font-size: 0.9rem;
        opacity: 0.8;
    }
    
    /* SUCESSO BOX */
    .success-box {
        background: linear-gradient(135deg, #11998e 0%, #38ef7d 100%);
        padding: 2rem;
        border-radius: 20px;
        color: white;
        text-align: center;
        margin: 2rem 0;
        box-shadow: 0 15px 35px rgba(17, 153, 142, 0.4);
    }
    
    .success-title {
        font-size: 2rem;
        font-weight: bold;
        margin-bottom: 1rem;
    }
    
    /* BOTÕES */
    .stButton>button {
        width: 100%;
        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
        color: white;
        font-size: 1.3rem;
        padding: 1rem 2rem;
        border-radius: 50px;
        border: none;
        font-weight: bold;
        box-shadow: 0 10px 30px rgba(102, 126, 234, 0.4);
        transition: all 0.3s ease;
        text-transform: uppercase;
        letter-spacing: 1px;
    }
    
    .stButton>button:hover {
        transform: translateY(-3px);
        box-shadow: 0 15px 40px rgba(102, 126, 234, 0.6);
    }
    
    .download-btn>button {
        background: linear-gradient(135deg, #f093fb 0%, #f5576c 100%);
        font-size: 1.5rem;
        padding: 1.5rem 3rem;
    }
    
    .reset-btn>button {
        background: linear-gradient(135deg, #4facfe 0%, #00f2fe 100%);
        font-size: 1rem;
        padding: 0.8rem 1.5rem;
    }
    
    /* INPUT FIELDS */
    .stTextInput>div>div>input {
        border-radius: 10px;
        border: 2px solid #e0e0e0;
        padding: 0.8rem;
        font-size: 1rem;
        transition: all 0.3s ease;
    }
    
    .stTextInput>div>div>input:focus {
        border-color: #667eea;
        box-shadow: 0 0 0 3px rgba(102, 126, 234, 0.1);
    }
    
    /* FILE UPLOADER */
    .uploadedFile {
        background: white;
        border-radius: 10px;
        padding: 1rem;
        margin: 0.5rem 0;
        box-shadow: 0 5px 15px rgba(0,0,0,0.1);
    }
    
    /* ANIMAÇÕES */
    @keyframes fadeInDown {
        from {
            opacity: 0;
            transform: translateY(-30px);
        }
        to {
            opacity: 1;
            transform: translateY(0);
        }
    }
    
    @keyframes bounce {
        0%, 100% {
            transform: translateY(0);
        }
        50% {
            transform: translateY(-10px);
        }
    }
    
    @keyframes pulse {
        0%, 100% {
            opacity: 1;
        }
        50% {
            opacity: 0.7;
        }
    }
    
    /* PROGRESS BAR */
    .stProgress > div > div > div {
        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
    }
    
    /* SIDEBAR */
    .css-1d391kg {
        background: linear-gradient(180deg, #667eea 0%, #764ba2 100%);
    }
    
    section[data-testid="stSidebar"] {
        background: linear-gradient(180deg, #667eea 0%, #764ba2 100%);
    }
    
    section[data-testid="stSidebar"] * {
        color: white !important;
    }
    
    /* FOOTER */
    .footer {
        text-align: center;
        padding: 2rem;
        background: rgba(255,255,255,0.95);
        border-radius: 20px;
        margin-top: 3rem;
        box-shadow: 0 10px 30px rgba(0,0,0,0.1);
    }
    
    .footer-brand {
        font-size: 1.3rem;
        font-weight: bold;
        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
        -webkit-background-clip: text;
        -webkit-text-fill-color: transparent;
        margin-bottom: 0.5rem;
    }
</style>
""", unsafe_allow_html=True)

# ═══════════════════════════════════════════════════════════════
# API KEY
# ═══════════════════════════════════════════════════════════════

ANTHROPIC_API_KEY = os.environ.get('ANTHROPIC_API_KEY')
if not ANTHROPIC_API_KEY:
    st.error("❌ ERRO: Variável ANTHROPIC_API_KEY não configurada!")
    st.info("Configure em Settings > Secrets no Streamlit Cloud")
    st.stop()

client = anthropic.Anthropic(api_key=ANTHROPIC_API_KEY)

# ═══════════════════════════════════════════════════════════════
# PROMPT V3.4 (MESMO DA V3.3.3)
# ═══════════════════════════════════════════════════════════════

PROMPT_ANALISE = """Você é um especialista em análise de extratos bancários e de aplicativos de mobilidade/delivery para aprovação de crédito imobiliário.

# TAREFA
Analise o extrato e extraia TODAS as entradas de dinheiro (créditos/receitas), excluindo transferências entre membros da mesma família.

# TIPOS DE FONTE DE RENDA VÁLIDOS

## BANCOS TRADICIONAIS:
• PIX recebido
• TED recebido
• DOC recebido
• Transferência recebida
• Depósito
• Crédito em conta
• Salário
• Pagamento
• Rendimento

## APLICATIVOS DE MOBILIDADE:
• Ganhos Uber
• Ganhos 99
• Ganhos da 99
• Receita Uber
• Receita 99
• Transferência Uber
• Pagamento Uber

## APLICATIVOS DE DELIVERY:
• Ganhos iFood
• Ganhos Rappi
• Receita iFood
• Receita Rappi
• Pagamento iFood
• Pagamento Rappi

## BANCOS DIGITAIS:
• Entrada Digio
• Entrada PicPay
• Entrada Mercado Pago
• Recebido via Pix (qualquer banco)
• Crédito (qualquer banco digital)

# AGRUPAMENTO POR DIA (CRÍTICO PARA APPS!)

**PARA APPS DE MOBILIDADE E DELIVERY (Uber/99/iFood/Rappi):**
- AGRUPE todas as entradas do MESMO DIA
- Some os valores do dia
- Crie UMA ÚNICA entrada por dia com o total
- Descrição: "Ganhos [App] - Total do dia"

**PARA BANCOS (tradicionais e digitais):**
- Liste cada transação separadamente (não agrupe)

# FILTRO DE FAMÍLIA (CRÍTICO!)

## REGRAS:
1. Compare o SOBRENOME do titular com sobrenomes nas descrições
2. EXCLUA transferências entre pessoas com mesmo sobrenome
3. Para APPS: NÃO aplique filtro de família
4. Para BANCOS: SEMPRE aplique filtro

# DETECÇÃO DE TIPO DE FONTE

Identifique automaticamente se o extrato é de:
- BANCO_TRADICIONAL: Caixa, BB, Itaú, Bradesco, Santander, Sicoob, Sicredi
- BANCO_DIGITAL: Nubank, Inter, Digio, PicPay, Mercado Pago
- APP_MOBILIDADE: Uber, 99
- APP_DELIVERY: iFood, Rappi

# FORMATO DA RESPOSTA (JSON VÁLIDO)

{
  "tipo_fonte": "BANCO_TRADICIONAL|BANCO_DIGITAL|APP_MOBILIDADE|APP_DELIVERY",
  "entradas": [
    {
      "data": "DD/MM/AAAA",
      "descricao": "Para apps: 'Ganhos [App] - Total do dia' | Para bancos: descrição original",
      "valor": 1234.56
    }
  ],
  "resumo": {
    "total_entradas": 10,
    "valor_total": 12345.67,
    "maior_entrada": 5000.00,
    "menor_entrada": 50.00,
    "media_mensal": 4115.22
  },
  "observacoes": "Qualquer informação relevante sobre o extrato"
}

CRÍTICO:
- Retorne APENAS o JSON
- JSON COMPLETO
- Apps: 1 entrada por dia
- Bancos: 1 entrada por transação
"""

# ═══════════════════════════════════════════════════════════════
# FUNÇÕES AUXILIARES (MESMAS DA V3.3.3)
# ═══════════════════════════════════════════════════════════════

def extrair_texto_pdf(pdf_bytes):
    """Extrai texto de PDF"""
    try:
        pdf_reader = PyPDF2.PdfReader(io.BytesIO(pdf_bytes))
        texto = ""
        for page in pdf_reader.pages:
            texto += page.extract_text() + "\n"
        return texto
    except Exception as e:
        return f"Erro ao extrair PDF: {str(e)}"

def processar_imagem(image_bytes):
    """Converte imagem para base64"""
    try:
        image = Image.open(io.BytesIO(image_bytes))
        if image.mode != 'RGB':
            image = image.convert('RGB')
        if image.width > 2000 or image.height > 2000:
            image.thumbnail((2000, 2000))
        img_byte_arr = io.BytesIO()
        image.save(img_byte_arr, format='JPEG')
        img_byte_arr = img_byte_arr.getvalue()
        return base64.b64encode(img_byte_arr).decode('utf-8')
    except Exception as e:
        return None

def analisar_com_claude(conteudo, tipo_arquivo, nome_cliente, banco):
    """Analisa extrato com Claude"""
    
    mensagem_contexto = f"""
CONTEXTO DA ANÁLISE:
• Cliente: {nome_cliente}
• Banco/App: {banco}
• Tipo de arquivo: {tipo_arquivo}

{PROMPT_ANALISE}
"""
    
    if tipo_arquivo in ['image/jpeg', 'image/png', 'image/jpg', 'image/webp']:
        content = [
            {
                "type": "image",
                "source": {
                    "type": "base64",
                    "media_type": tipo_arquivo,
                    "data": conteudo
                }
            },
            {
                "type": "text",
                "text": mensagem_contexto
            }
        ]
    else:
        content = [
            {
                "type": "text",
                "text": f"{mensagem_contexto}\n\nEXTRATO:\n{conteudo}"
            }
        ]
    
    try:
        response_text = ""
        
        with client.messages.stream(
            model="claude-sonnet-4-20250514",
            max_tokens=32000,
            temperature=0,
            messages=[{
                "role": "user",
                "content": content
            }]
        ) as stream:
            for text in stream.text_stream:
                response_text += text
        
        return response_text
        
    except Exception as e:
        return f"Erro na API: {str(e)}"

def completar_json_parcial(json_parcial, entradas_encontradas):
    """Completa JSON parcial"""
    try:
        if 'resumo' in json_parcial and json_parcial['resumo']:
            return json_parcial
        
        if entradas_encontradas and len(entradas_encontradas) > 0:
            valores = [e.get('valor', 0) for e in entradas_encontradas if 'valor' in e]
            
            if valores:
                resumo = {
                    'total_entradas': len(entradas_encontradas),
                    'valor_total': sum(valores),
                    'maior_entrada': max(valores),
                    'menor_entrada': min(valores),
                    'media_mensal': sum(valores) / 3 if len(valores) > 0 else 0
                }
                
                json_parcial['resumo'] = resumo
        
        if 'observacoes' not in json_parcial:
            json_parcial['observacoes'] = 'Análise concluída com sucesso'
        
        return json_parcial
        
    except Exception as e:
        return json_parcial

def validar_e_corrigir_json(texto_resposta):
    """Valida e corrige JSON"""
    
    texto = texto_resposta.strip()
    
    if texto.startswith("```json"):
        texto = texto[7:]
    if texto.startswith("```"):
        texto = texto[3:]
    if texto.endswith("```"):
        texto = texto[:-3]
    texto = texto.strip()
    
    # Camada 1: Parse direto
    try:
        dados = json.loads(texto)
        return dados, None
    except:
        pass
    
    # Camada 2: Regex
    try:
        matches = re.finditer(r'\{(?:[^{}]|(?:\{[^{}]*\}))*\}', texto, re.DOTALL)
        for match in matches:
            try:
                dados = json.loads(match.group())
                if 'entradas' in dados:
                    return dados, None
            except:
                continue
    except:
        pass
    
    # Camada 3: Reconstrução
    try:
        tipo_match = re.search(r'"tipo_fonte"\s*:\s*"([^"]+)"', texto)
        entradas_match = re.search(r'"entradas"\s*:\s*\[(.*?)\]', texto, re.DOTALL)
        
        if tipo_match and entradas_match:
            tipo_fonte = tipo_match.group(1)
            entradas_str = entradas_match.group(1)
            
            entradas = []
            entrada_pattern = r'\{[^}]+\}'
            for entrada_match in re.finditer(entrada_pattern, entradas_str):
                try:
                    entrada = json.loads(entrada_match.group())
                    if 'data' in entrada and 'valor' in entrada:
                        entradas.append(entrada)
                except:
                    continue
            
            if entradas:
                dados_parciais = {
                    'tipo_fonte': tipo_fonte,
                    'entradas': entradas
                }
                
                dados_completos = completar_json_parcial(dados_parciais, entradas)
                return dados_completos, None
                
    except:
        pass
    
    return None, f"Não foi possível extrair JSON válido"

def criar_excel_profissional(dados_json, nome_cliente, banco):
    """Cria Excel profissional"""
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Análise de Renda"
    
    # Estilos
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    subheader_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    subheader_font = Font(bold=True, size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Cabeçalho
    ws.merge_cells('A1:D1')
    ws['A1'] = "ANÁLISE DE RENDA BANCÁRIA"
    ws['A1'].font = Font(bold=True, size=16, color="FFFFFF")
    ws['A1'].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30
    
    # Info cliente
    ws['A3'] = "CLIENTE:"
    ws['B3'] = nome_cliente
    ws['A4'] = "BANCO/APP:"
    ws['B4'] = banco
    ws['A5'] = "DATA ANÁLISE:"
    ws['B5'] = datetime.now().strftime("%d/%m/%Y %H:%M")
    ws['A6'] = "TIPO FONTE:"
    ws['B6'] = dados_json.get('tipo_fonte', 'N/A')
    
    for row in [3, 4, 5, 6]:
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'].font = Font(color="1f77b4")
    
    # Resumo
    ws['A8'] = "RESUMO FINANCEIRO"
    ws.merge_cells('A8:D8')
    ws['A8'].font = subheader_font
    ws['A8'].fill = subheader_fill
    ws['A8'].alignment = Alignment(horizontal='center')
    
    resumo = dados_json.get('resumo', {})
    ws['A9'] = "Total de Entradas:"
    ws['B9'] = resumo.get('total_entradas', 0)
    ws['A10'] = "Valor Total:"
    ws['B10'] = f"R$ {resumo.get('valor_total', 0):,.2f}"
    ws['A11'] = "Maior Entrada:"
    ws['B11'] = f"R$ {resumo.get('maior_entrada', 0):,.2f}"
    ws['A12'] = "Menor Entrada:"
    ws['B12'] = f"R$ {resumo.get('menor_entrada', 0):,.2f}"
    ws['A13'] = "Média Mensal:"
    ws['B13'] = f"R$ {resumo.get('media_mensal', 0):,.2f}"
    
    for row in range(9, 14):
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'].font = Font(color="28a745", size=11)
    
    # Tabela de entradas
    ws['A15'] = "DETALHAMENTO DAS ENTRADAS"
    ws.merge_cells('A15:D15')
    ws['A15'].font = subheader_font
    ws['A15'].fill = subheader_fill
    ws['A15'].alignment = Alignment(horizontal='center')
    
    # Headers
    headers = ['DATA', 'DESCRIÇÃO', 'VALOR', 'STATUS']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=16, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # Dados
    entradas = dados_json.get('entradas', [])
    for idx, entrada in enumerate(entradas, start=17):
        ws[f'A{idx}'] = entrada.get('data', '')
        ws[f'B{idx}'] = entrada.get('descricao', '')
        ws[f'C{idx}'] = f"R$ {entrada.get('valor', 0):,.2f}"
        ws[f'D{idx}'] = "✓ Válido"
        
        ws[f'C{idx}'].font = Font(color="28a745", bold=True)
        ws[f'D{idx}'].font = Font(color="28a745")
        
        for col in ['A', 'B', 'C', 'D']:
            ws[f'{col}{idx}'].border = border
            ws[f'{col}{idx}'].alignment = Alignment(vertical='center')
    
    # Observações
    if 'observacoes' in dados_json:
        obs_row = len(entradas) + 18
        ws[f'A{obs_row}'] = "OBSERVAÇÕES:"
        ws[f'A{obs_row}'].font = Font(bold=True)
        ws[f'B{obs_row}'] = dados_json['observacoes']
        ws.merge_cells(f'B{obs_row}:D{obs_row}')
    
    # Ajusta larguras
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 50
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 12
    
    # Salva
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output

def criar_grafico_entradas(entradas):
    """Cria gráfico de barras das entradas por dia"""
    
    # Agrupa por data
    valores_por_data = defaultdict(float)
    for entrada in entradas:
        data = entrada.get('data', '')
        valor = entrada.get('valor', 0)
        valores_por_data[data] += valor
    
    # Ordena por data
    datas = sorted(valores_por_data.keys())
    valores = [valores_por_data[d] for d in datas]
    
    # Pega últimos 30 dias
    if len(datas) > 30:
        datas = datas[-30:]
        valores = valores[-30:]
    
    # Cria gráfico
    fig = go.Figure()
    
    fig.add_trace(go.Bar(
        x=datas,
        y=valores,
        marker=dict(
            color=valores,
            colorscale='Viridis',
            showscale=False
        ),
        text=[f'R$ {v:,.2f}' for v in valores],
        textposition='outside',
        hovertemplate='<b>%{x}</b><br>R$ %{y:,.2f}<extra></extra>'
    ))
    
    fig.update_layout(
        title={
            'text': '💰 Entradas por Dia (Últimos 30 dias)',
            'x': 0.5,
            'xanchor': 'center',
            'font': {'size': 20, 'color': '#333', 'family': 'Arial'}
        },
        xaxis_title='Data',
        yaxis_title='Valor (R$)',
        plot_bgcolor='rgba(0,0,0,0)',
        paper_bgcolor='rgba(0,0,0,0)',
        font=dict(size=12),
        height=400,
        showlegend=False,
        hovermode='x unified'
    )
    
    fig.update_xaxes(showgrid=True, gridwidth=1, gridcolor='rgba(200,200,200,0.2)')
    fig.update_yaxes(showgrid=True, gridwidth=1, gridcolor='rgba(200,200,200,0.2)')
    
    return fig

# ═══════════════════════════════════════════════════════════════
# INTERFACE STREAMLIT
# ═══════════════════════════════════════════════════════════════

def main():
    
    # HERO HEADER
    st.markdown("""
    <div class="hero-header">
        <div class="hero-icon">💰</div>
        <h1 class="hero-title">Analisador de Renda Pro</h1>
        <p class="hero-subtitle">Análise Profissional de Extratos Bancários com IA</p>
    </div>
    """, unsafe_allow_html=True)
    
    # Sidebar
    with st.sidebar:
        st.markdown("### ℹ️ Sobre o Sistema")
        st.info("""
        **V3.4 - Design Profissional**
        
        ✅ Interface moderna
        ✅ Dashboard visual
        ✅ Gráficos interativos
        ✅ 15 tipos suportados
        """)
        
        st.markdown("### 📋 Tipos Suportados")
        
        with st.expander("🏦 Bancos Tradicionais"):
            st.markdown("""
            • Caixa • BB • Itaú
            • Bradesco • Santander
            • Sicoob • Sicredi
            """)
        
        with st.expander("💳 Bancos Digitais"):
            st.markdown("""
            • Nubank • Inter
            • Digio • PicPay
            • Mercado Pago
            """)
        
        with st.expander("🚗 Apps Mobilidade"):
            st.markdown("• Uber • 99")
        
        with st.expander("🍔 Apps Delivery"):
            st.markdown("• iFood • Rappi")
    
    # FEATURES CARDS
    col1, col2, col3 = st.columns(3)
    
    with col1:
        st.markdown("""
        <div class="feature-card">
            <div class="feature-icon">🤖</div>
            <div class="feature-title">IA Avançada</div>
            <div class="feature-desc">Claude Sonnet 4 analisa seus extratos com precisão</div>
        </div>
        """, unsafe_allow_html=True)
    
    with col2:
        st.markdown("""
        <div class="feature-card">
            <div class="feature-icon">📊</div>
            <div class="feature-title">Dashboard Visual</div>
            <div class="feature-desc">Visualize suas finanças com gráficos profissionais</div>
        </div>
        """, unsafe_allow_html=True)
    
    with col3:
        st.markdown("""
        <div class="feature-card">
            <div class="feature-icon">⚡</div>
            <div class="feature-title">Super Rápido</div>
            <div class="feature-desc">Processamento em tempo real com streaming</div>
        </div>
        """, unsafe_allow_html=True)
    
    st.markdown("<br>", unsafe_allow_html=True)
    
    # INPUTS
    col1, col2 = st.columns(2)
    
    with col1:
        nome_cliente = st.text_input(
            "👤 Nome Completo do Cliente",
            placeholder="Ex: João Silva Santos"
        )
    
    with col2:
        banco = st.text_input(
            "🏦 Banco ou App",
            placeholder="Ex: Caixa, Uber, iFood"
        )
    
    st.markdown("### 📤 Upload de Extratos")
    
    uploaded_files = st.file_uploader(
        "Arraste os arquivos ou clique para selecionar",
        type=['pdf', 'png', 'jpg', 'jpeg'],
        accept_multiple_files=True,
        help="Formatos: PDF, PNG, JPG, JPEG"
    )
    
    if uploaded_files:
        st.success(f"✅ {len(uploaded_files)} arquivo(s) selecionado(s)")
        for file in uploaded_files:
            st.text(f"📄 {file.name} ({file.size / 1024:.1f} KB)")
    
    st.markdown("<br>", unsafe_allow_html=True)
    
    # BOTÕES
    col_btn1, col_btn2 = st.columns([3, 1])
    
    with col_btn1:
        processar = st.button("🚀 PROCESSAR EXTRATOS", type="primary")
    
    with col_btn2:
        st.markdown('<div class="reset-btn">', unsafe_allow_html=True)
        if st.button("🔄 NOVA CONSULTA"):
            st.rerun()
        st.markdown('</div>', unsafe_allow_html=True)
    
    if processar:
        
        # Validações
        if not nome_cliente or not banco:
            st.error("⚠️ Preencha o nome do cliente e o banco!")
            return
        
        if not uploaded_files:
            st.error("⚠️ Faça upload de pelo menos um extrato!")
            return
        
        # Processamento
        st.markdown("---")
        progress_bar = st.progress(0)
        status_text = st.empty()
        
        todas_entradas = []
        erros = []
        
        for idx, file in enumerate(uploaded_files):
            
            status_text.text(f"📄 Processando: {file.name}...")
            progress = (idx + 1) / len(uploaded_files)
            progress_bar.progress(progress)
            
            try:
                file_bytes = file.read()
                file_type = file.type
                
                if file_type == 'application/pdf':
                    conteudo = extrair_texto_pdf(file_bytes)
                elif file_type in ['image/jpeg', 'image/png', 'image/jpg']:
                    conteudo = processar_imagem(file_bytes)
                else:
                    erros.append(f"{file.name}: Tipo não suportado")
                    continue
                
                status_text.text(f"🤖 Analisando {file.name} com IA...")
                resposta = analisar_com_claude(conteudo, file_type, nome_cliente, banco)
                
                dados, erro = validar_e_corrigir_json(resposta)
                
                if erro:
                    erros.append(f"{file.name}: {erro}")
                    continue
                
                if dados and 'entradas' in dados:
                    todas_entradas.extend(dados['entradas'])
            
            except Exception as e:
                erros.append(f"{file.name}: {str(e)}")
        
        progress_bar.progress(1.0)
        status_text.empty()
        
        # RESULTADOS
        if todas_entradas:
            
            # Calcula resumo
            valores = [e['valor'] for e in todas_entradas]
            resumo = {
                'total_entradas': len(todas_entradas),
                'valor_total': sum(valores),
                'maior_entrada': max(valores),
                'menor_entrada': min(valores),
                'media_mensal': sum(valores) / 3 if len(valores) > 0 else 0
            }
            
            dados_completos = {
                'tipo_fonte': 'MÚLTIPLOS_ARQUIVOS',
                'entradas': todas_entradas,
                'resumo': resumo,
                'observacoes': f'Análise de {len(uploaded_files)} arquivo(s)'
            }
            
            # SUCCESS BOX
            st.markdown("""
            <div class="success-box">
                <div class="success-title">✅ Análise Concluída com Sucesso!</div>
                <p>Seus dados foram processados e estão prontos para visualização</p>
            </div>
            """, unsafe_allow_html=True)
            
            # DASHBOARD DE MÉTRICAS
            st.markdown("## 📊 Dashboard Financeiro")
            
            col1, col2, col3 = st.columns(3)
            
            with col1:
                st.markdown(f"""
                <div class="metric-card">
                    <div class="metric-icon">💰</div>
                    <div class="metric-label">Valor Total</div>
                    <div class="metric-value">R$ {resumo['valor_total']:,.2f}</div>
                    <div class="metric-subtitle">Soma de todas as entradas</div>
                </div>
                """, unsafe_allow_html=True)
            
            with col2:
                st.markdown(f"""
                <div class="metric-card">
                    <div class="metric-icon">📈</div>
                    <div class="metric-label">Média Mensal</div>
                    <div class="metric-value">R$ {resumo['media_mensal']:,.2f}</div>
                    <div class="metric-subtitle">Baseado em 3 meses</div>
                </div>
                """, unsafe_allow_html=True)
            
            with col3:
                st.markdown(f"""
                <div class="metric-card">
                    <div class="metric-icon">📝</div>
                    <div class="metric-label">Total Entradas</div>
                    <div class="metric-value">{resumo['total_entradas']}</div>
                    <div class="metric-subtitle">Transações processadas</div>
                </div>
                """, unsafe_allow_html=True)
            
            st.markdown("<br><br>", unsafe_allow_html=True)
            
            # GRÁFICO
            fig = criar_grafico_entradas(todas_entradas)
            st.plotly_chart(fig, use_container_width=True)
            
            st.markdown("<br>", unsafe_allow_html=True)
            
            # DOWNLOAD EXCEL
            excel_file = criar_excel_profissional(dados_completos, nome_cliente, banco)
            
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"analise_renda_{nome_cliente.replace(' ', '_')}_{timestamp}.xlsx"
            
            col_download, col_reset = st.columns([3, 1])
            
            with col_download:
                st.markdown('<div class="download-btn">', unsafe_allow_html=True)
                st.download_button(
                    label="📊 BAIXAR RELATÓRIO COMPLETO",
                    data=excel_file,
                    file_name=filename,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
                st.markdown('</div>', unsafe_allow_html=True)
            
            with col_reset:
                st.markdown('<div class="reset-btn">', unsafe_allow_html=True)
                if st.button("🔄 NOVA ANÁLISE", key="reset2"):
                    st.rerun()
                st.markdown('</div>', unsafe_allow_html=True)
        
        # Erros
        if erros:
            st.markdown("---")
            st.markdown("### ⚠️ Avisos")
            for erro in erros:
                st.warning(erro)
    
    # FOOTER
    st.markdown("""
    <div class="footer">
        <p class="footer-brand">By Magalhães Negócios</p>
        <p><strong>Analisador de Renda V3.4</strong> | Design Profissional</p>
        <p style='font-size: 0.9rem; color: #666; margin-top: 0.5rem;'>
            Desenvolvido com ❤️ usando Streamlit + Claude Sonnet 4
        </p>
        <p style='font-size: 0.8rem; color: #999; margin-top: 0.3rem;'>
            Powered by Anthropic AI
        </p>
    </div>
    """, unsafe_allow_html=True)

if __name__ == "__main__":
    main()
